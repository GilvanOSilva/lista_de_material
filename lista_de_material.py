import PySimpleGUI as Gui
from openpyxl import load_workbook
from random import  randint
senha = 'admin5678'


def template_do_menu(modo):
    return [
        [Gui.Button(modo), Gui.Button('Voltar'), Gui.Text('', key='info')],
        [Gui.Text('Produto', size=35),
         Gui.Text('Código', size=7),
         Gui.Text('Preço', size=7),
         Gui.Text('Quantidade', size=8)
         ]
    ]


def trocar(variavel):
    variavel = str(variavel)
    if ',' in variavel:
        variavel = variavel.replace(',', '.')
        return float(variavel)
    else:
        variavel = variavel.replace('.', ',')
        return str(variavel)


def numero_do_pedido():
    numero = ''
    for i in range(6):
        numero += str(randint(0, 9))
    return numero


def modo_pedido():
    file_xl = load_workbook(filename='produtos.xlsx')
    produtos = file_xl.active
    posicao_tabela = 1
    pedido = template_do_menu('Gerar Pedido')
    while produtos['A' + str(posicao_tabela)].value is not None:
        pedido.append([
            Gui.Text(produtos['A' + str(posicao_tabela)].value, size=35),
            Gui.Text(produtos['B' + str(posicao_tabela)].value, size=7),
            Gui.Text(produtos['C' + str(posicao_tabela)].value, size=7),
            Gui.Input('0', key='pedido_quantidade' + str(posicao_tabela), size=8),
            Gui.Checkbox(text='', default=False, key=posicao_tabela)
            ])
        posicao_tabela += 1
    janela_pedido = Gui.Window('Gerar Pedido', pedido)
    while True:
        eventos, valores = janela_pedido.read()
        if eventos == 'Gerar Pedido':
            lista = []
            posicao_tabela = 1
            while produtos['A' + str(posicao_tabela)].value is not None:
                if int(valores['pedido_quantidade' + str(posicao_tabela)]) > int(produtos['D' + str(posicao_tabela)].value):
                    # janela_pedido['info'].update(produtos['A' + str(posicao_tabela)].value + ' indisponível, solicite até: ' + produtos['D' + str(posicao_tabela)].value)
                    break
                if valores[posicao_tabela] is True:
                    lista.append(
                        [produtos['A' + str(posicao_tabela)].value,
                         int(valores['pedido_quantidade' + str(posicao_tabela)]),
                         produtos['C' + str(posicao_tabela)].value,
                         int(valores['pedido_quantidade' + str(posicao_tabela)]) * produtos['C' + str(posicao_tabela)].value
                         ])
                    posicao_tabela += 1
                else:
                    posicao_tabela += 1
            janela_pedido.hide()
            finalizado = modo_finalizar(lista)
            if finalizado is False:
                janela_pedido.UnHide()
            else:
                posicao_tabela = 1
                while produtos['A' + str(posicao_tabela)].value is not None:
                    if valores[posicao_tabela] is True:
                        janela_pedido['pedido_quantidade' + str(posicao_tabela)].update('0')
                        janela_pedido[posicao_tabela].update(False)
                        posicao_tabela += 1
                    else:
                        posicao_tabela += 1
                janela_pedido.UnHide()

        if eventos == Gui.WINDOW_CLOSED or eventos == 'Voltar':
            janela_pedido.close()
            break


def modo_finalizar(lista):
    finalizar = [
        [Gui.Button('Finalizar Pedido'), Gui.Button('Voltar'), Gui.Text('', key='info')],
        [Gui.Text('Produto', size=35),
         Gui.Text('Quantidade', size=7),
         Gui.Text('Preço', size=7),
         Gui.Text('Total', size=7),
         ]
    ]
    i = 0
    total_produtos = 0
    total_preco = 0
    while i < len(lista):
        total_produtos += lista[i][1]
        total_preco += lista[i][3]
        finalizar.append([
            Gui.Text(lista[i][0], size=35),
            Gui.Text(lista[i][1], size=7),
            Gui.Text(trocar(lista[i][2]), size=7),
            Gui.Text(trocar(lista[i][3]), size=7)
        ])
        i += 1
    total_preco = trocar(total_preco)
    finalizar.append([
        [Gui.Text('=' * 60)],
        Gui.Text('Total do pedido', size=35),
        Gui.Text(total_produtos, size=16),
        Gui.Text(total_preco, size=5),
    ])
    janela_finalizar = Gui.Window('Finalizar Pedido', finalizar)
    while True:
        eventos, valores = janela_finalizar.read()
        if eventos == 'Finalizar Pedido':
            if total_preco == '0':
                janela_finalizar['info'].update('Seu pedido está vazio.')
            else:
                numero = numero_do_pedido()
                txtfile = open(numero + '.txt', 'x')
                txtfile = open(numero + '.txt', 'w')
                txtfile.write('Número do pedido: ' + numero + '\n')
                for i in range(len(lista)):
                    txtfile.write(lista[i][0] + ' x ' + str(lista[i][1]) + '---' + trocar(lista[i][2]) + '---' + trocar(lista[i][3]) + '\n')
                txtfile.write('Total de produtos: ' + str(total_produtos) + '\n')
                txtfile.write('Valor total: ' + total_preco + '\n')
                txtfile.close()
                atualizar_lista(lista, numero)
                janela_finalizar.close()
                return True
        if eventos == Gui.WINDOW_CLOSED or eventos == 'Voltar':
            janela_finalizar.close()
            return False


def atualizar_lista(lista, numero):
    file_xl = load_workbook(filename='produtos.xlsx')
    produtos = file_xl.active
    posicao_tabela = 1
    posicao_pedido = 0
    variavel = [
        [Gui.Text('Seu pedido foi realizado com sucesso!', size=40)],
        [Gui.Text('Número do pedido:' + numero, size=40)],
        [Gui.Text('Informe seu "Número de pedido" no local de retirada.', size=40)],
        [Gui.Button('Pedido Completo', size=40)]
         ]
    janela_variavel = Gui.Window('Pedido completo', variavel)
    while produtos['A' + str(posicao_tabela)].value is not None and posicao_pedido < len(lista):
        if produtos['A' + str(posicao_tabela)].value == lista[posicao_pedido][0]:
            produtos['D' + str(posicao_tabela)].value = produtos['D' + str(posicao_tabela)].value - lista[posicao_pedido][1]
            posicao_tabela += 1
            posicao_pedido += 1
        else:
            posicao_tabela += 1
        file_xl.save(filename='produtos.xlsx')
    while True:
        eventos, valores = janela_variavel.read()
        if eventos == Gui.WINDOW_CLOSED or eventos == 'Pedido Completo':
            janela_variavel.close()
            break


def modo_manutencao():
    manutencao = [
        [Gui.Text('Modo Editor')],
        [Gui.Button('Adicionar', size=9), Gui.Button('Remover', size=9)],
        [Gui.Button('Alterar', size=9), Gui.Button('Voltar', size=9)]
    ]
    janela_manutencao = Gui.Window('Editor', manutencao)
    janela_principal.Hide()
    while True:
        eventos, valores = janela_manutencao.read()
        if eventos == 'Adicionar':
            janela_manutencao.Hide()
            modo_adicao()
            janela_manutencao.UnHide()
        if eventos == 'Remover':
            janela_manutencao.Hide()
            modo_remocao()
            janela_manutencao.UnHide()
        if eventos == 'Alterar':
            janela_manutencao.Hide()
            modo_alteracao()
            janela_manutencao.UnHide()
        if eventos == Gui.WINDOW_CLOSED or eventos == 'Voltar':
            janela_principal.UnHide()
            janela_manutencao.close()
            break


def modo_adicao():
    file_xl = load_workbook(filename='produtos.xlsx')
    produtos = file_xl.active
    posicao_tabela = 1
    adicao = template_do_menu('Adicionar produto')
    adicao.append([Gui.Input(key='adicao_produto', size=40),
                   Gui.Input(key='adicao_codigo', size=8),
                   Gui.Input(key='adicao_preco', size=8),
                   Gui.Input(key='adicao_quantidade', size=8)
                   ])
    janela_adicao = Gui.Window('Adicionar produto', adicao)
    while True:
        eventos, valores = janela_adicao.read()
        if eventos == 'Adicionar produto':
            while produtos['A' + str(posicao_tabela)].value is not None:
                posicao_tabela += 1
            produtos['A' + str(posicao_tabela)] = str(valores['adicao_produto'])
            produtos['B' + str(posicao_tabela)] = int(valores['adicao_codigo'])
            produtos['C' + str(posicao_tabela)] = str(valores['adicao_preco'])
            produtos['D' + str(posicao_tabela)] = int(valores['adicao_quantidade'])
            file_xl.save(filename='produtos.xlsx')
            janela_adicao['info'].update('Produto adicionado.')
            janela_adicao['adicao_produto'].update('')
            janela_adicao['adicao_codigo'].update('')
            janela_adicao['adicao_preco'].update('')
            janela_adicao['adicao_quantidade'].update('')
        if eventos == Gui.WINDOW_CLOSED or eventos == 'Voltar':
            janela_adicao.close()
            break


def modo_alteracao():
    file_xl = load_workbook(filename='produtos.xlsx')
    produtos = file_xl.active
    posicao_tabela = 1
    alteracao = template_do_menu('Alterar produto')
    while produtos['A' + str(posicao_tabela)].value is not None:
        alteracao.append([
         Gui.Input(produtos['A' + str(posicao_tabela)].value, key='alterar_produto' + str(posicao_tabela), size=40),
         Gui.Input(produtos['B' + str(posicao_tabela)].value, key='alterar_codigo' + str(posicao_tabela), size=8),
         Gui.Input(trocar(produtos['C' + str(posicao_tabela)].value), key='alterar_preco' + str(posicao_tabela), size=8),
         Gui.Input(produtos['D' + str(posicao_tabela)].value, key='alterar_quantidade' + str(posicao_tabela), size=8),
         Gui.Checkbox(text='', default=False, key=posicao_tabela)
                          ])
        posicao_tabela += 1
    janela_alteracao = Gui.Window('Alterar produto', alteracao)
    while True:
        eventos, valores = janela_alteracao.read()
        if eventos == 'Alterar produto':
            posicao_tabela = 1
            while produtos['A' + str(posicao_tabela)].value is not None:
                if valores[posicao_tabela] is True:
                    produtos['A' + str(posicao_tabela)].value = str(valores['alterar_produto' + str(posicao_tabela)])
                    produtos['B' + str(posicao_tabela)].value = int(valores['alterar_codigo' + str(posicao_tabela)])
                    produtos['C' + str(posicao_tabela)].value = trocar(valores['alterar_preco' + str(posicao_tabela)])
                    produtos['D' + str(posicao_tabela)].value = int(valores['alterar_quantidade' + str(posicao_tabela)])
                    file_xl.save(filename='produtos.xlsx')
                    janela_alteracao['info'].update('O produto selecionado alterado.')
                    janela_alteracao[posicao_tabela].update(False)
                    posicao_tabela += 1
                else:
                    posicao_tabela += 1
        if eventos == Gui.WINDOW_CLOSED or eventos == 'Voltar':
            janela_alteracao.close()
            break


def modo_remocao():
    file_xl = load_workbook(filename='produtos.xlsx')
    produtos = file_xl.active
    posicao_tabela = 1
    remocao = template_do_menu('Remover produto')
    while produtos['A' + str(posicao_tabela)].value is not None:
        remocao.append([Gui.Text(produtos['A' + str(posicao_tabela)].value, size=35),
                        Gui.Text(produtos['B' + str(posicao_tabela)].value, size=7),
                        Gui.Text(produtos['C' + str(posicao_tabela)].value, size=7),
                        Gui.Text(produtos['D' + str(posicao_tabela)].value, size=7),
                        Gui.Checkbox(text='', default=False, key=posicao_tabela)
                        ])
        posicao_tabela += 1
    janela_remocao = Gui.Window('Remover produto', remocao)
    while True:
        eventos, valores = janela_remocao.read()
        if eventos == 'Remover produto':
            posicao_tabela = 1
            limpar = 0
            cache = []
            while produtos['A' + str(posicao_tabela)].value is not None:
                if valores[posicao_tabela] is True:
                    limpar += 1
                    posicao_tabela += 1
                else:
                    cache.append([
                         produtos['A' + str(posicao_tabela)].value, produtos['B' + str(posicao_tabela)].value,
                         produtos['C' + str(posicao_tabela)].value, produtos['D' + str(posicao_tabela)].value
                         ])
                    posicao_tabela += 1
            posicao_tabela = 1
            for i in range(len(cache)):
                produtos['A' + str(posicao_tabela)].value = cache[posicao_tabela - 1][0]
                produtos['B' + str(posicao_tabela)].value = cache[posicao_tabela - 1][1]
                produtos['C' + str(posicao_tabela)].value = cache[posicao_tabela - 1][2]
                produtos['D' + str(posicao_tabela)].value = cache[posicao_tabela - 1][3]
                posicao_tabela += 1
            for i in range(limpar):
                produtos['A' + str(posicao_tabela)].value = None
                produtos['B' + str(posicao_tabela)].value = None
                produtos['C' + str(posicao_tabela)].value = None
                produtos['D' + str(posicao_tabela)].value = None
                posicao_tabela += 1
            file_xl.save(filename='produtos.xlsx')
            posicao_tabela = 1
            remocao = template_do_menu('Remover produto')
            while produtos['A' + str(posicao_tabela)].value is not None:
                remocao.append([Gui.Text(produtos['A' + str(posicao_tabela)].value, size=35),
                                Gui.Text(produtos['B' + str(posicao_tabela)].value, size=7),
                                Gui.Text(produtos['C' + str(posicao_tabela)].value, size=7),
                                Gui.Text(produtos['D' + str(posicao_tabela)].value, size=7),
                                Gui.Checkbox(text='', default=False, key=posicao_tabela)
                                ])
                posicao_tabela += 1
            janela_remocao.close()
            janela_remocao = Gui.Window('Remover produto', remocao)
        if eventos == Gui.WINDOW_CLOSED or eventos == 'Voltar':
            janela_remocao.close()
            break


principal = [
    [Gui.Text('Estoque da papelaria')],
    [Gui.Button('Geração de pedido', size=15)],
    [Gui.Button('Manutenção da lista', size=15), Gui.Input(key='senha', password_char='*', size=10, do_not_clear=False)],
    [Gui.Button('Sair', size=15)]
]
janela_principal = Gui.Window('Lista de material', principal)
while True:
    eventos, valores = janela_principal.read()
    valores['senha'] = 'admin5678' # remover linha após teste
    if eventos == 'Geração de pedido':
        janela_principal.Hide()
        modo_pedido()
        janela_principal.UnHide()
    if eventos == 'Manutenção da lista' and valores['senha'] == senha:
        janela_principal.Hide()
        modo_manutencao()
        janela_principal.UnHide()
    if eventos == Gui.WINDOW_CLOSED or eventos == 'Sair':
        break
